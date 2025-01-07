from openpyxl import Workbook, load_workbook


# Implementing the list object
class Node:
    def __init__(self, d, n):
        self.data = d
        self.next = n

class LinkedList:
    def __init__(self):
        self.head = None
        self.length = 0

    def __str__(self):
        if self.head == None: 
            return "empty"
        st = "-"
        ptr = self.head
        while ptr != None:
            
            st += "-> "+str(ptr.data)+" "
            ptr = ptr.next
        return st+"|"
        
    def search(self, d):
        i = 0
        ptr = self.head
        while ptr != None:
            if ptr.data == d:
                return i
            ptr = ptr.next
            i += 1
        return None
    
    #ONLY FOR SHELVES
    def searchByShelfID(self, id):
        i = 0
        ptr = self.head
        while ptr != None:
            if str(ptr.data.id) == str(id):
                return ptr.data
            ptr = ptr.next
            i += 1
        return None
    
    def filterByShelfID(self, list, id):
        i = 0
        ptr = list.head
        newList = LinkedList()
        if (list.head == None):
            return None
        else:
            while ptr != None:
                if ptr.data.id == id:
                    newList.append(ptr.data)
                ptr = ptr.next
        return newList
    
    # only for shoes
    def searchByUPC(self, UPC):
        i = 0
        ptr = self.head
        while ptr != None:
            if str(ptr.data.UPC) == str(UPC):
                return ptr.data
            ptr = ptr.next
            i += 1
        return None
    
    def append(self, d):
        if self.head == None:      
            self.head = Node(d,None) 
        else:
            ptr = self.head
            while ptr.next != None:
                ptr = ptr.next
            ptr.next = Node(d,None)
        self.length += 1

    def insert(self, i, d):
        if self.head == None or i == 0:
            self.head = Node(d,self.head)
        else:
            ptr = self.head
            while i>1 and ptr.next != None:
                ptr = ptr.next
                i -= 1
            ptr.next = Node(d,ptr.next)
        self.length += 1

    def remove(self, i): # removes i-th element and returns it
        if self.head == None:
            return None
        if i == 0:
            val = self.head.data
            self.head = self.head.next
            self.length -= 1
            return val
        ptr = self.head
        while ptr.next != None:
            if i == 1:
                val = ptr.next.data
                ptr.next = ptr.next.next
                self.length -= 1
                return val                
            ptr = ptr.next
            i -= 1
            
    def removeVal(self, d):
        if self.head == None:
            return -1
        if self.head.data == d:
            self.head = self.head.next
            self.length -= 1
            return 0
        else:
            i = 0
            ptr = self.head	
            while ptr.next != None:
                if ptr.next.data == d:
                    ptr.next = ptr.next.next
                    self.length -= 1
                    return i+1
                ptr = ptr.next
                i += 1
        return -1
    
    def get(self, i):
        ptr = self.head
        while i > 0:
            ptr = ptr.next
            i -= 1
        return ptr.data
    
    def set(self, i, d):
        ptr = self.head
        while i > 0:
            ptr = ptr.next
            i -= 1
        ptr.data = d

# Shoe object
class shoe:
    def __init__(self, StyleID, name, cat, minSize, maxSize, UPC, size):
        self.StyleID = StyleID
        self.name = name
        self.minSize = minSize
        self.maxSize = maxSize
        self.cat = cat
        self.UPC = UPC
        self.size = size

# Shelf object
class shelf:
    def __init__(self, id):
        self.id = id
        self.count = 0
        self.contents = LinkedList()
    
    def addToShelf(self,shoe):
        self.contents.append(shoe)
        self.count += 1

class gap:
    def __init__(self, shoe, shelf):
        self.name = shoe.name
        self.StyleID = shoe.StyleID
        self.size = shoe.size
        self.id = shelf
        self.status = "PENDING"

class db:
    def __init__(self):
        self.shoesList = LinkedList()
        self.shelvesList = LinkedList()
        self.gapRequests = LinkedList()
        self.wb = load_workbook(filename = 'db.xlsx')
        self.shoeSheet = self.wb['Sheet1']
        self.shelfSheet = self.wb['Sheet2']
        
        i = 2
        while(self.shoeSheet.cell(row=i,column=1).value != None):
            self.shoesList.append(shoe(
                                  self.shoeSheet.cell(row=i,column=1).value,
                                  self.shoeSheet.cell(row=i,column=2).value,
                                  self.shoeSheet.cell(row=i,column=3).value,
                                  self.shoeSheet.cell(row=i,column=4).value,
                                  self.shoeSheet.cell(row=i,column=5).value,
                                  self.shoeSheet.cell(row=i,column=6).value,
                                  self.shoeSheet.cell(row=i,column=7).value))
            
            i += 1

        i = 2
        while(self.shelfSheet.cell(row=i,column=1).value != None):
            temp = shelf(self.shelfSheet.cell(row=i,column=1).value)

            x = 2
            while(self.shelfSheet.cell(row=i,column=x).value != None):
                UPC = self.shelfSheet.cell(row=i,column=x).value
                temp.addToShelf(self.shoesList.searchByUPC(UPC))
                x += 1

            self.shelvesList.append(temp)
            i += 1



    def searchShelvesForShoe(self,shoe):
        noOfShelves = self.shelvesList.length
        for i in range (0,noOfShelves):
            thisShelf = self.shelvesList.get(i)
            print(thisShelf.id)
            print(thisShelf.contents.get(0))

            if (thisShelf.contents.search(shoe)) != None:
                return str(thisShelf.id)
        return None

class system:
    def __init__(self):
        self.db = db()
        self.home()

    def home(self):
        print("1. Footwear Fill Up")
        print("2. Gap Requests")
        print("3. Exit")
        op = input("Select line:")

        while(True):
            if op == "1":
                self.fillUp()
                break
            elif op == "2":
                self.gaps()
                break
            elif op == "3":
                return
            else:
                print("Enter valid input!")
                self.home()

    def printGapInfo(self, gap):
        print(gap.StyleID)
        print(gap.size)
        print(gap.id)
        print(gap.status)

    def listGaps(self):
        l = self.db.gapRequests.length
        list = self.db.gapRequests

        print("______________________")
        print("Gap Requests")
        print("  ")
        for i in range (0,l):
            thatGap = list.get(i)
            print(str(i+1) + ": " + thatGap.name)
            self.printGapInfo(thatGap)
        print("______________________")
        print("  ")

    def listGapsOnShelf(self,shelf):
        gapsList = self.db.gapRequests
        newlist = self.db.gapRequests.filterByShelfID(gapsList, shelf)
        
        l = newlist.length

        if l == 0:
            print ("No Gaps on Shelf!")
        else:
            print("______________________")
            print("Gap Requests on shelf" + " " + str(shelf))
            for i in range (0,l):
                thisGap = newlist.get(i)
                
                print(str(i+1) + ": " + thisGap.name)
                print(thisGap.StyleID)
                print(thisGap.size)
                print(thisGap.id)
                print(thisGap.status)
            print("______________________")
            print("  ")

        return newlist

    def gaps(self):
        l = self.db.gapRequests.length
        list = self.db.gapRequests

        self.listGaps()
        
        while True:
            op = input("Scan shelf or 'B' to return:")
            if (op in ("B" , "b")) == False:
                shelfGapsList = self.listGapsOnShelf(op)
            else:
                self.home()
                return
            
            if shelfGapsList.length > 0:
                gapSelect = shelfGapsList.get(int(input("Select ln:")) - 1)
                print(" ")
                print(gapSelect.name)
                self.printGapInfo(gapSelect)
                print(" ")
                conf = input("1: Pick 2: Cancel 3: Missing B: Return")
                if conf == ("1"):
                    print("PICKED")
                    gapSelect.status = "PICKED"
                if conf == ("2"):
                    print("CANCELLED")
                    gapSelect.status = "CANCELLED"
                if conf == ("3"):
                    print("MISSING")
                    gapSelect.status = "MISSING"
                if conf in ("B" , "b"):
                    self.gaps()
                else:
                    print("INVALID")
                self.gaps()
                break

    def fillUp(self):
        UPC = str(input("Scan Barcode or 'B' to return"))
        if (UPC in ("B" , "b")) == False:
            shoe = self.db.shoesList.searchByUPC(UPC)
            if (shoe != None):
                conf = input("Confirm print Y/N")
                if conf == ("y" or "Y"):
                    fill = gap(shoe, self.db.searchShelvesForShoe(shoe))
                    self.db.gapRequests.append(fill)
                    print("Label Printed")
                    self.home()
                else:
                    self.fillUp()
            else:
                print("No such shoe")
                self.home()
        else:
            self.home()
    
def main():
    app = system()

main()